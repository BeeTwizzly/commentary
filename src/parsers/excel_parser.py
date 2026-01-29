"""Excel parser for Brinson attribution workbooks."""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models import HoldingData, ParsedWorkbook, StrategyHoldings

logger = logging.getLogger(__name__)

# Column name variations we should recognize (case-insensitive)
COLUMN_PATTERNS = {
    "ticker": [r"ticker", r"symbol", r"security\s*id"],
    "company_name": [r"company", r"name", r"security\s*name", r"issuer"],
    "avg_weight": [r"avg\.?\s*weight", r"average\s*weight", r"avg\s*wt"],
    "begin_weight": [r"begin\.?\s*weight", r"start\s*weight", r"beg\.?\s*wt"],
    "end_weight": [r"end\.?\s*weight", r"ending\s*weight", r"end\s*wt"],
    "benchmark_weight": [r"bench\.?\s*weight", r"benchmark\s*wt", r"bmk\s*wt"],
    "benchmark_return": [r"bench\.?\s*return", r"benchmark\s*ret", r"bmk\s*ret"],
    "total_attribution": [r"total\s*attribution", r"total\s*attrib", r"attribution", r"total\s*effect"],
    "selection_effect": [r"selection\s*effect", r"selection", r"stock\s*selection"],
    "allocation_effect": [r"allocation\s*effect", r"allocation", r"asset\s*allocation"],
}


def parse_workbook(file: BinaryIO | Path | str) -> ParsedWorkbook:
    """
    Parse an Excel workbook containing Brinson attribution data.

    Args:
        file: File path, Path object, or file-like object (e.g., from Streamlit uploader)

    Returns:
        ParsedWorkbook containing all strategies with top/bottom 5 holdings

    Raises:
        ValueError: If the file cannot be parsed as an Excel workbook
    """
    result = ParsedWorkbook()

    try:
        wb = load_workbook(filename=file, read_only=True, data_only=True)
    except Exception as e:
        result.errors.append(f"Failed to open workbook: {e}")
        logger.error("Failed to open workbook: %s", e)
        return result

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        strategy_result = _parse_strategy_sheet(sheet, sheet_name)

        if strategy_result.all_holdings:
            result.strategies.append(strategy_result)
            logger.info(
                "Parsed strategy '%s': %d holdings, %d contributors, %d detractors",
                sheet_name,
                len(strategy_result.all_holdings),
                len(strategy_result.contributors),
                len(strategy_result.detractors),
            )
        else:
            logger.warning("Sheet '%s' contained no valid holdings", sheet_name)

    wb.close()
    return result


def _parse_strategy_sheet(sheet: Worksheet, strategy_name: str) -> StrategyHoldings:
    """Parse a single strategy sheet and extract top/bottom 5 holdings."""
    result = StrategyHoldings(strategy_name=strategy_name)

    # Find header row and column mapping
    column_map, header_row = _find_columns(sheet)

    if not column_map or "ticker" not in column_map:
        result.parse_warnings.append(f"Could not identify column structure in '{strategy_name}'")
        return result

    # Parse data rows
    holdings: list[HoldingData] = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        holding = _parse_row(row, column_map, strategy_name, row_idx, result.parse_warnings)
        if holding:
            holdings.append(holding)

    if not holdings:
        return result

    # Sort by attribution and assign ranks
    holdings_sorted = sorted(holdings, key=lambda h: h.total_attribution, reverse=True)

    # Create ranked holdings
    ranked_holdings = []
    for rank, h in enumerate(holdings_sorted, start=1):
        ranked_holdings.append(HoldingData(
            ticker=h.ticker,
            company_name=h.company_name,
            strategy=h.strategy,
            avg_weight=h.avg_weight,
            begin_weight=h.begin_weight,
            end_weight=h.end_weight,
            benchmark_weight=h.benchmark_weight,
            benchmark_return=h.benchmark_return,
            total_attribution=h.total_attribution,
            selection_effect=h.selection_effect,
            allocation_effect=h.allocation_effect,
            rank=rank,
            is_contributor=rank <= 5,
        ))

    result.all_holdings = ranked_holdings
    result.contributors = [h for h in ranked_holdings if h.rank <= 5][:5]
    result.detractors = [h for h in ranked_holdings if h.rank > len(ranked_holdings) - 5][-5:]
    # Re-sort detractors so worst is first
    result.detractors = sorted(result.detractors, key=lambda h: h.total_attribution)

    return result


def _find_columns(sheet: Worksheet) -> tuple[dict[str, int], int]:
    """
    Scan sheet to find header row and map column names to indices.

    Returns:
        Tuple of (column_map, header_row_number)
        column_map keys are our standard names, values are 0-based column indices
    """
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
        column_map = {}

        for col_idx, cell_value in enumerate(row):
            if cell_value is None:
                continue

            cell_str = str(cell_value).strip().lower()

            for standard_name, patterns in COLUMN_PATTERNS.items():
                if standard_name in column_map:
                    continue
                for pattern in patterns:
                    if re.search(pattern, cell_str, re.IGNORECASE):
                        column_map[standard_name] = col_idx
                        break

        # If we found ticker and at least total_attribution, we have a valid header
        if "ticker" in column_map and "total_attribution" in column_map:
            return column_map, row_idx

    return {}, 0


def _parse_row(
    row: tuple,
    column_map: dict[str, int],
    strategy_name: str,
    row_idx: int,
    warnings: list[str],
) -> HoldingData | None:
    """Parse a single data row into a HoldingData object."""

    def get_value(key: str, default=None):
        if key not in column_map:
            return default
        idx = column_map[key]
        if idx >= len(row):
            return default
        return row[idx]

    def get_float(key: str, default: float = 0.0) -> float:
        val = get_value(key)
        if val is None:
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    ticker = get_value("ticker")
    company_name = get_value("company_name", "")

    # Skip rows without valid ticker
    if not ticker or not isinstance(ticker, str):
        return None

    ticker = ticker.strip().upper()

    # Skip likely summary/total rows
    if _is_summary_row(ticker, company_name):
        return None

    # Validate ticker format (basic check)
    if not re.match(r"^[A-Z]{1,5}(\.[A-Z])?(/[A-Z])?$", ticker):
        warnings.append(f"Row {row_idx}: Skipped invalid ticker format '{ticker}'")
        return None

    total_attribution = get_float("total_attribution")

    # Skip rows with zero attribution (likely empty or invalid)
    if total_attribution == 0.0:
        return None

    return HoldingData(
        ticker=ticker,
        company_name=str(company_name).strip() if company_name else "",
        strategy=strategy_name,
        avg_weight=get_float("avg_weight"),
        begin_weight=get_float("begin_weight"),
        end_weight=get_float("end_weight"),
        benchmark_weight=get_float("benchmark_weight"),
        benchmark_return=get_float("benchmark_return"),
        total_attribution=total_attribution,
        selection_effect=get_float("selection_effect"),
        allocation_effect=get_float("allocation_effect"),
        rank=0,  # Will be set after sorting
        is_contributor=False,  # Will be set after sorting
    )


def _is_summary_row(ticker: str, company_name: str) -> bool:
    """Check if this row is likely a summary/total row to skip."""
    summary_patterns = [
        r"^total",
        r"^sum",
        r"^grand\s*total",
        r"^portfolio",
        r"^benchmark",
        r"^cash",
        r"^residual",
    ]

    combined = f"{ticker} {company_name}".lower()

    for pattern in summary_patterns:
        if re.search(pattern, combined):
            return True

    return False
