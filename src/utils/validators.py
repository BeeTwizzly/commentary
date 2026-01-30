"""Input validation utilities."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO

import openpyxl

logger = logging.getLogger(__name__)

# Constants
MAX_FILE_SIZE_MB = 50
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
MIN_ROWS_REQUIRED = 2  # Header + at least 1 data row
MAX_ROWS_ALLOWED = 10000

# Required columns (case-insensitive matching)
REQUIRED_COLUMNS = {"ticker", "total effect"}


@dataclass
class ValidationResult:
    """Result of a validation check."""

    valid: bool
    error_message: str = ""
    warnings: list[str] | None = None

    def __post_init__(self):
        if self.warnings is None:
            self.warnings = []


def validate_excel_file(
    file: BinaryIO,
    filename: str,
) -> ValidationResult:
    """
    Validate an uploaded Excel file.

    Args:
        file: File-like object with Excel data
        filename: Original filename

    Returns:
        ValidationResult with status and any errors/warnings
    """
    warnings = []

    # Check extension
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return ValidationResult(
            valid=False,
            error_message=f"Invalid file type '{ext}'. Please upload an Excel file (.xlsx or .xls).",
        )

    # Check file size
    file.seek(0, 2)  # Seek to end
    size_bytes = file.tell()
    file.seek(0)  # Reset to beginning

    size_mb = size_bytes / (1024 * 1024)
    if size_mb > MAX_FILE_SIZE_MB:
        return ValidationResult(
            valid=False,
            error_message=f"File too large ({size_mb:.1f} MB). Maximum allowed is {MAX_FILE_SIZE_MB} MB.",
        )

    if size_mb > MAX_FILE_SIZE_MB * 0.8:
        warnings.append(f"Large file ({size_mb:.1f} MB) may take longer to process.")

    # Try to open workbook
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Failed to open Excel file: %s", e)
        return ValidationResult(
            valid=False,
            error_message="Could not read Excel file. Please ensure it's a valid .xlsx file and not password-protected.",
        )

    # Check for sheets
    if not wb.sheetnames:
        return ValidationResult(
            valid=False,
            error_message="Excel file contains no worksheets.",
        )

    # Check first sheet has data
    sheet = wb.active
    if sheet is None:
        return ValidationResult(
            valid=False,
            error_message="Could not access worksheet.",
        )

    # Check row count
    row_count = sheet.max_row or 0
    if row_count < MIN_ROWS_REQUIRED:
        return ValidationResult(
            valid=False,
            error_message=f"File appears empty or has only headers. Found {row_count} rows.",
        )

    if row_count > MAX_ROWS_ALLOWED:
        return ValidationResult(
            valid=False,
            error_message=f"File has too many rows ({row_count:,}). Maximum allowed is {MAX_ROWS_ALLOWED:,}.",
        )

    # Check for required columns
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).lower().strip())

    headers_set = set(headers)
    missing_required = []

    for req in REQUIRED_COLUMNS:
        # Fuzzy match - check if required term appears in any header
        found = any(req in h for h in headers_set)
        if not found:
            missing_required.append(req)

    if missing_required:
        return ValidationResult(
            valid=False,
            error_message=f"Missing required columns: {', '.join(missing_required)}. Found columns: {', '.join(headers[:10])}{'...' if len(headers) > 10 else ''}",
        )

    wb.close()

    logger.info(
        "Excel validation passed: %s (%d rows, %.1f MB)",
        filename,
        row_count,
        size_mb,
    )

    return ValidationResult(valid=True, warnings=warnings)


def validate_ticker(ticker: str) -> ValidationResult:
    """
    Validate a stock ticker symbol.

    Args:
        ticker: Ticker symbol to validate

    Returns:
        ValidationResult
    """
    if not ticker:
        return ValidationResult(valid=False, error_message="Ticker cannot be empty")

    ticker = ticker.strip().upper()

    # Basic format check: 1-5 alphanumeric characters
    if not re.match(r"^[A-Z]{1,5}$", ticker):
        # Allow for special tickers like BRK.A, BRK.B
        if not re.match(r"^[A-Z]{1,5}\.[A-Z]$", ticker):
            return ValidationResult(
                valid=False,
                error_message=f"Invalid ticker format: '{ticker}'. Expected 1-5 letters (e.g., AAPL, MSFT).",
            )

    return ValidationResult(valid=True)


def validate_strategy_name(name: str) -> ValidationResult:
    """
    Validate a strategy name.

    Args:
        name: Strategy name to validate

    Returns:
        ValidationResult
    """
    if not name:
        return ValidationResult(valid=False, error_message="Strategy name cannot be empty")

    name = name.strip()

    if len(name) < 2:
        return ValidationResult(
            valid=False,
            error_message="Strategy name too short (minimum 2 characters).",
        )

    if len(name) > 100:
        return ValidationResult(
            valid=False,
            error_message="Strategy name too long (maximum 100 characters).",
        )

    return ValidationResult(valid=True)


def validate_quarter(quarter: str) -> ValidationResult:
    """
    Validate a quarter string.

    Args:
        quarter: Quarter string (e.g., "Q4 2025")

    Returns:
        ValidationResult
    """
    if not quarter:
        return ValidationResult(valid=False, error_message="Quarter cannot be empty")

    quarter = quarter.strip()

    # Expected format: Q1-Q4 followed by 4-digit year
    pattern = r"^Q[1-4]\s*20[2-3][0-9]$"
    if not re.match(pattern, quarter, re.IGNORECASE):
        return ValidationResult(
            valid=False,
            error_message=f"Invalid quarter format: '{quarter}'. Expected format: 'Q4 2025'.",
        )

    return ValidationResult(valid=True)


def validate_api_key(key: str) -> ValidationResult:
    """
    Validate OpenAI API key format (not actual validity).

    Args:
        key: API key string

    Returns:
        ValidationResult
    """
    if not key:
        return ValidationResult(
            valid=False,
            error_message="API key is required. Set OPENAI_API_KEY in environment or .streamlit/secrets.toml",
        )

    key = key.strip()

    # OpenAI keys start with sk- and are ~50 chars
    if not key.startswith("sk-"):
        return ValidationResult(
            valid=False,
            error_message="Invalid API key format. OpenAI keys should start with 'sk-'.",
        )

    if len(key) < 20:
        return ValidationResult(
            valid=False,
            error_message="API key appears too short.",
        )

    return ValidationResult(valid=True)
