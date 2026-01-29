"""Tests for the Excel parser module."""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.parsers.excel_parser import parse_workbook, _find_columns, _is_summary_row
from src.models import HoldingData, ParsedWorkbook


@pytest.fixture
def sample_workbook(tmp_path) -> Path:
    """Create a sample Excel workbook for testing."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create "Large Cap Growth" strategy sheet
    ws1 = wb.create_sheet("Large Cap Growth")
    ws1.append(["Ticker", "Company Name", "Avg Weight", "Begin Weight", "End Weight",
                "Benchmark Weight", "Benchmark Return", "Total Attribution",
                "Selection Effect", "Allocation Effect"])

    # Add 15 holdings with varying attribution
    holdings_data = [
        ("NVDA", "NVIDIA Corp", 4.2, 3.8, 4.6, 3.1, 12.5, 45.0, 32.0, 13.0),
        ("AAPL", "Apple Inc", 5.1, 5.0, 5.2, 6.2, 8.3, 35.0, 25.0, 10.0),
        ("MSFT", "Microsoft Corp", 6.0, 5.8, 6.2, 5.5, 10.0, 28.0, 20.0, 8.0),
        ("GOOGL", "Alphabet Inc", 3.5, 3.4, 3.6, 4.0, 7.5, 22.0, 18.0, 4.0),
        ("AMZN", "Amazon.com Inc", 4.0, 3.9, 4.1, 3.8, 9.0, 18.0, 12.0, 6.0),
        ("META", "Meta Platforms", 2.8, 2.7, 2.9, 2.5, 6.0, 12.0, 8.0, 4.0),
        ("TSLA", "Tesla Inc", 2.0, 2.2, 1.8, 1.5, -5.0, 5.0, 3.0, 2.0),
        ("JPM", "JPMorgan Chase", 3.0, 3.0, 3.0, 3.2, 4.0, 2.0, 1.0, 1.0),
        ("V", "Visa Inc", 2.5, 2.5, 2.5, 2.8, 3.0, -5.0, -3.0, -2.0),
        ("JNJ", "Johnson & Johnson", 2.2, 2.2, 2.2, 2.5, 1.0, -12.0, -8.0, -4.0),
        ("PG", "Procter & Gamble", 1.8, 1.8, 1.8, 2.0, 0.5, -18.0, -12.0, -6.0),
        ("KO", "Coca-Cola Co", 1.5, 1.5, 1.5, 1.8, -1.0, -25.0, -18.0, -7.0),
        ("PEP", "PepsiCo Inc", 1.6, 1.6, 1.6, 1.9, -2.0, -30.0, -22.0, -8.0),
        ("WMT", "Walmart Inc", 1.4, 1.4, 1.4, 1.7, -3.0, -38.0, -28.0, -10.0),
        ("XOM", "Exxon Mobil", 1.2, 1.2, 1.2, 1.5, -4.0, -42.0, -30.0, -12.0),
    ]

    for row in holdings_data:
        ws1.append(row)

    # Add a summary row that should be skipped
    ws1.append(["TOTAL", "Portfolio Total", 100.0, 100.0, 100.0, 100.0, 5.0, 0.0, 0.0, 0.0])

    # Create "Small Cap Value" strategy sheet
    ws2 = wb.create_sheet("Small Cap Value")
    ws2.append(["Ticker", "Company Name", "Avg Weight", "Begin Weight", "End Weight",
                "Benchmark Weight", "Benchmark Return", "Total Attribution",
                "Selection Effect", "Allocation Effect"])

    small_cap_data = [
        ("ABC", "ABC Corp", 2.0, 1.9, 2.1, 1.5, 15.0, 50.0, 35.0, 15.0),
        ("DEF", "DEF Inc", 1.8, 1.7, 1.9, 1.4, 12.0, 40.0, 28.0, 12.0),
        ("GHI", "GHI Ltd", 1.5, 1.4, 1.6, 1.2, 10.0, 30.0, 20.0, 10.0),
        ("JKL", "JKL Corp", 1.3, 1.2, 1.4, 1.0, 8.0, 20.0, 14.0, 6.0),
        ("MNO", "MNO Inc", 1.2, 1.1, 1.3, 0.9, 6.0, 15.0, 10.0, 5.0),
        ("PQR", "PQR Ltd", 1.0, 0.9, 1.1, 0.8, 4.0, 5.0, 3.0, 2.0),
        ("STU", "STU Corp", 0.9, 0.8, 1.0, 0.7, 2.0, -10.0, -7.0, -3.0),
        ("VWX", "VWX Inc", 0.8, 0.7, 0.9, 0.6, -2.0, -20.0, -14.0, -6.0),
        ("YZA", "YZA Ltd", 0.7, 0.6, 0.8, 0.5, -4.0, -30.0, -22.0, -8.0),
        ("BCD", "BCD Corp", 0.6, 0.5, 0.7, 0.4, -6.0, -40.0, -28.0, -12.0),
    ]

    for row in small_cap_data:
        ws2.append(row)

    # Save workbook
    file_path = tmp_path / "test_performance.xlsx"
    wb.save(file_path)
    wb.close()

    return file_path


class TestParseWorkbook:
    """Tests for the main parse_workbook function."""

    def test_parses_multiple_strategies(self, sample_workbook):
        """Should parse all strategy tabs in workbook."""
        result = parse_workbook(sample_workbook)

        assert len(result.strategies) == 2
        assert result.strategies[0].strategy_name == "Large Cap Growth"
        assert result.strategies[1].strategy_name == "Small Cap Value"

    def test_identifies_top_5_contributors(self, sample_workbook):
        """Should correctly identify top 5 by attribution."""
        result = parse_workbook(sample_workbook)
        lcg = result.strategies[0]

        assert len(lcg.contributors) == 5

        # Top 5 should be NVDA, AAPL, MSFT, GOOGL, AMZN (highest attribution)
        contributor_tickers = [h.ticker for h in lcg.contributors]
        assert contributor_tickers == ["NVDA", "AAPL", "MSFT", "GOOGL", "AMZN"]

        # All should be marked as contributors
        assert all(h.is_contributor for h in lcg.contributors)

        # Ranks should be 1-5
        assert [h.rank for h in lcg.contributors] == [1, 2, 3, 4, 5]

    def test_identifies_bottom_5_detractors(self, sample_workbook):
        """Should correctly identify bottom 5 by attribution."""
        result = parse_workbook(sample_workbook)
        lcg = result.strategies[0]

        assert len(lcg.detractors) == 5

        # Bottom 5 should be XOM, WMT, PEP, KO, PG (lowest attribution, sorted worst first)
        detractor_tickers = [h.ticker for h in lcg.detractors]
        assert detractor_tickers == ["XOM", "WMT", "PEP", "KO", "PG"]

        # None should be marked as contributors
        assert not any(h.is_contributor for h in lcg.detractors)

    def test_skips_summary_rows(self, sample_workbook):
        """Should skip rows that look like totals/summaries."""
        result = parse_workbook(sample_workbook)
        lcg = result.strategies[0]

        # Should have 15 holdings, not 16 (TOTAL row skipped)
        assert len(lcg.all_holdings) == 15
        assert not any(h.ticker == "TOTAL" for h in lcg.all_holdings)

    def test_returns_total_counts(self, sample_workbook):
        """Should provide accurate total counts."""
        result = parse_workbook(sample_workbook)

        assert result.total_holdings == 25  # 15 + 10
        assert result.total_for_commentary == 20  # (5+5) * 2 strategies

    def test_get_all_commentary_holdings(self, sample_workbook):
        """Should flatten all holdings for commentary."""
        result = parse_workbook(sample_workbook)
        holdings = result.get_all_commentary_holdings()

        assert len(holdings) == 20

    def test_handles_missing_file(self):
        """Should return error for missing file."""
        result = parse_workbook("/nonexistent/path.xlsx")

        assert len(result.errors) > 0
        assert len(result.strategies) == 0

    def test_attribution_description(self, sample_workbook):
        """Should generate human-readable attribution description."""
        result = parse_workbook(sample_workbook)
        top = result.strategies[0].contributors[0]
        bottom = result.strategies[0].detractors[0]

        assert "added" in top.attribution_description()
        assert "45" in top.attribution_description()
        assert "detracted" in bottom.attribution_description()


class TestFindColumns:
    """Tests for column detection logic."""

    def test_finds_standard_headers(self, sample_workbook):
        """Should find columns with standard naming."""
        from openpyxl import load_workbook
        wb = load_workbook(sample_workbook, read_only=True)
        sheet = wb["Large Cap Growth"]

        column_map, header_row = _find_columns(sheet)

        assert "ticker" in column_map
        assert "company_name" in column_map
        assert "total_attribution" in column_map
        assert header_row == 1

        wb.close()


class TestIsSummaryRow:
    """Tests for summary row detection."""

    def test_detects_total_row(self):
        assert _is_summary_row("TOTAL", "Portfolio Total")
        assert _is_summary_row("total", "")

    def test_detects_cash_row(self):
        assert _is_summary_row("CASH", "Cash & Equivalents")

    def test_allows_valid_tickers(self):
        assert not _is_summary_row("NVDA", "NVIDIA Corp")
        assert not _is_summary_row("AAPL", "Apple Inc")
